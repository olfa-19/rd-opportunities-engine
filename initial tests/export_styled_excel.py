# export_styled_excel.py
import os
import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.join(os.getcwd(), "나라장터")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "G2B_Executive_Report.xlsx")

def export_styled_excel():
    records = []
    for folder in os.listdir(BASE_DIR):
        meta_path = os.path.join(BASE_DIR, folder, "metadata.json")
        if os.path.isdir(os.path.join(BASE_DIR, folder)) and os.path.exists(meta_path):
            with open(meta_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                files = [a["file_name"] for a in data.get("attachments", [])]
                records.append({
                    "Bid ID": data.get("bid_id", ""),
                    "Posting Date": data.get("posting_date", ""),
                    "Organization": data.get("organization", ""),
                    "Opportunity Title": data.get("title", ""),
                    "File Count": len(files),
                    "Scraped Attachments": "\n".join(files) if files else "None",
                    "G2B Link": data.get("url", "")
                })

    df = pd.DataFrame(records)
    if df.empty:
        return

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="R&D Feed", startrow=4)
        ws = writer.sheets["R&D Feed"]

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        kpi_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Title Block & KPI Header
        ws["A1"] = "R&D Opportunity Engine — Executive Summary"
        ws["A1"].font = title_font
        ws["A2"] = f"Total Tracked Bids: {len(df)}   |   Total Files Extracted: {df['File Count'].sum()}"
        ws["A2"].font = kpi_font

        # Format Data Headers
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=5, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format Data Rows
        for row in range(6, len(df) + 6):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Highlight row if files exist
                if col == 5 and cell.value > 0:
                    cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")

        # Column Widths
        widths = [16, 14, 25, 45, 12, 35, 30]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    print(f"[+] Styled Excel saved to {OUTPUT_EXCEL}")

if __name__ == "__main__":
    export_styled_excel()